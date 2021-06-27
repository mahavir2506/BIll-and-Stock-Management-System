import sys
from PyQt5.QtWidgets import QApplication, QCheckBox, QMainWindow, QLabel, QComboBox, QPushButton,QCompleter,QWidget,QTabWidget,QVBoxLayout,QLineEdit,QTableWidget,QTableWidgetItem,QMessageBox
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont,QDoubleValidator,QIntValidator
from PyQt5.QtCore import Qt, QSortFilterProxyModel
import json
from openpyxl import load_workbook
import os
from xlsx2html import xlsx2html
from datetime import date
from openpyxl.styles import Alignment, Font
os.chdir(os.path.dirname(__file__))
path=os.getcwd()+"\\"

# Stock Detail Table Class
def refresh():
    bill_no=Bill().get_new_bill_no()
    ex.billing.c_name_text.clear()
    ex.billing.i_pricepergram_text.clear()
    ex.billing.received_i_pricepergram_text.clear()
    ex.billing.labor_charge_text.setText("0")
    ex.billing.c_name_text.setCurrentText('')
    ex.billing.c_mobileno_text.clear()
    ex.billing.c_address_text.clear()
    ex.billing.c_billno_text.clear()
    ex.billing.i_name_text.clear()
    ex.billing.i_weight_text.setText("0")
    ex.billing.i_price_text.setText("0")
    ex.billing.i_pricepergram_text.setText("0")
    ex.billing.item_table.setRowCount(0)
    ex.billing.received_i_name_text.clear()
    ex.billing.received_i_weight_text.setText("0")
    ex.billing.received_i_pricepergram_text.setText("0")
    ex.billing.received_i_price_text.setText("0")
    ex.billing.pan.setChecked(False)
    ex.billing.received_item_table.setRowCount(0)
    ex.billing.total_price=0.0
    ex.billing.overall_total=0.0
    ex.billing.pending_amount=0.0
    ex.billing.received_total_price=0.0
    ex.billing.pending_Ammount_value.setText('00000000000000')
    ex.billing.pending_Ammount_value.setHidden(True)
    ex.billing.total_Buying_Ammount_value.setText('00000000000000')
    ex.billing.total_Buying_Ammount_value.setHidden(True)
    ex.billing.collected_item_amount_value.setText('00000000000000')
    ex.billing.collected_item_amount_value.setHidden(True)
    ex.billing.total_amount_value.setText('00000000000000')
    ex.billing.total_amount_value.setHidden(True)
    ex.billing.collectd_amount_text.setText("0")
    #load detail from database
    ex.billing.customer_list.clear()
    c=Customer_detail()
    temp=c.get_all_data()
    for i  in temp:
        ex.billing.l.append(i[0])
        ex.billing.customer_list.append(i[0]["name"])
    ex.billing.customer_list=list(set(ex.billing.customer_list))
    ex.billing.c_name_text.addItems(ex.billing.customer_list)
    #ex.billing.c_billno_text.setText(str(int(d["bill_no"])+1))
    ex.billing.item_list.clear()
    print("fbdhbfdb step1")
    s=Stock_detail()
    temp=s.get_all_data()
    for i  in temp:
        ex.billing.l.append(i[0])
        ex.billing.item_list.append(i[0]["item_name"])
    ex.billing.item_list=list(set(ex.billing.item_list))
    ex.billing.i_name_text.addItems(ex.billing.item_list)
    ex.billing.i_name_text.setCurrentText('')
    ex.billing.c_name_text.setCurrentText('')
    ex.billing.item_validator_list.clear()
    ex.billing.item_list_load()
    ex.billing.received_i_name_text.setCurrentText('')
    ex.billing.c_address_text.setText("")
    ex.billing.c_mobileno_text.setText("")
    ex.billing.total_price=0
    ex.billing.overall_total=0
    ex.billing.pending_amount=0
    ex.billing.received_total_price=0
    ex.billing.pending_Ammount_value.setText('00000000000000')
    ex.billing.pending_Ammount_value.setHidden(True)
    ex.billing.total_Buying_Ammount_value.setText('00000000000000')
    ex.billing.total_Buying_Ammount_value.setHidden(True)
    ex.billing.collected_item_amount_value.setText('00000000000000')
    ex.billing.collected_item_amount_value.setHidden(True)
    ex.billing.total_amount_value.setText('00000000000000')
    ex.billing.total_amount_value.setHidden(True)
    ex.billing.c_billno_text.setText(str(bill_no))
    ex.billing.i_weight_text.setText("0")
    ex.billing.i_pricepergram_text.setText("0")
    print("bhfbsbfgb")
    c=Customer_detail()
    data=c.get_all_data()
    row = 0
    ex.customer_detail.customer_table.setRowCount(0)
    for i in data:
        ex.customer_detail.customer_table.setRowCount(row+1)
        l=[]
        l.append(i[0]["name"])
        l.append(i[0]["add"])
        l.append(i[0]["pending_amount"])
        l.append(i[0]["mob_no"])
        l.append(i[0]["bill no's list"])
        l.append(QPushButton("Delete",ex.customer_detail))
        l.append(QPushButton("Update",ex.customer_detail))
        col = 0
        for i in l:
            if col == 5:
                ex.customer_detail.customer_table.setCellWidget(row, 5, l[5])
                l[5].clicked.connect(ex.customer_detail.cellClick)
            elif col==6:
                ex.customer_detail.customer_table.setCellWidget(row, 6, l[6])
                l[6].clicked.connect(ex.customer_detail.update_data)
            else:
                cell = QTableWidgetItem(str(i))
                ex.customer_detail.customer_table.setItem(row, col, cell)
            col += 1
        row+=1
    c=Stock_detail()
    data=c.get_all_data()
    print(data)
    row = 0
    ex.stock_detail.stock_table.setRowCount(0)
    for i in data:
        ex.stock_detail.stock_table.setRowCount(row+1)
        l=[]
        l.append(i[0]["item_name"])
        l.append(i[0]["item_weight"])
        l.append(QPushButton("Delete",ex.stock_detail))
        l.append(QPushButton("Update",ex.stock_detail))
        col = 0
        for i in l:
            if col == 2:
                ex.stock_detail.stock_table.setCellWidget(row, 2, l[2])
                l[2].clicked.connect(ex.stock_detail.cellClick)
            elif col==3:
                ex.stock_detail.stock_table.setCellWidget(row, 3, l[3])
                l[3].clicked.connect(ex.stock_detail.update_data)
            else:
                cell = QTableWidgetItem(str(i))
                ex.stock_detail.stock_table.setItem(row, col, cell)
            col += 1
        row+=1
    ex.add_stock.i_name_text.clear()
    ex.add_stock.i_weight_text.clear()
    ex.add_stock.item_table.setRowCount(0)
    ex.add_stock.item_list.clear()
    s=Stock_detail()
    temp=s.get_all_data()
    for i  in temp:
        ex.add_stock.l.append(i[0])
        ex.add_stock.item_list.append(i[0]["item_name"])
    ex.add_stock.item_list=list(set(ex.add_stock.item_list))
    ex.add_stock.i_name_text.addItems(ex.add_stock.item_list)
    ex.add_stock.i_name_text.setCurrentText('')
    
    
class Stock_detail:
    #Insert Query
    def insert_data(self,name,weight):
        l=[]
        d={}
        found=0
        for i in self.data[0]["stock_detail"]:
            if i[0]["item_name"]==name:
                found=1
                i[0]["item_weight"]=str(float(i[0]["item_weight"])+float(weight))
        if found==0:
            d["item_name"]=name
            d["item_weight"]=weight
            l.append(d)
            self.data[0]["stock_detail"].append(l)
        f=open(path+"stock_detail.json",'w')
        f.write(json.dumps(self.data))
        f.close()

    
    def get_data(self,name):
        l=[]
        for i in self.data[0]["stock_detail"]:
            if name==i[0]["item_name"]:
                l.append(i[0])
        return l

    def get_all_data(self):
        return self.data[0]["stock_detail"]
    
    def delete_data(self,name):
        for i in self.data[0]["stock_detail"]:
            if i[0]["item_name"]==name:
                self.data[0]["stock_detail"].remove(i)
                f=open(path+'stock_detail.json','w')
                f.write(json.dumps(self.data))
                f.close()
                return True
    def update_data(self,name,weight):
        row=float(ex.stock_detail.stock_table.currentRow())
        for i in self.data[0]["stock_detail"]:
            if name==i[0]["item_name"]:
                i[0]["item_weight"]=ex.stock_detail.stock_table.item(row,1).text()
                f=open(path+'stock_detail.json','w')
                f.write(json.dumps(self.data))
                f.close()
                return True
        return True
    def __init__(self):
        f = open(path+"stock_detail.json")
        self.data = json.loads(f.read())
        f.close()
        
class Collected_item:
    def insert_data(self,l,bill_no):
        for i in l:
            l1=[]
            d4={}
            d4["bill_no"]=bill_no
            d4["received_item_name"]=i["received_item_name"]
            d4["received_item_weight"]=i["received_item_weight"]
            d4["received_item_pricepergram"]=i["received_item_pricepergram"]
            d4["received_item_price"]=i["received_item_price"]
            l1.append(d4)
            self.data[0]["Collected_item"].append(l1)
        f=open(path+'Collected_item.json','w')
        f.write(json.dumps(self.data))
        f.close()
    
    
    def __init__(self):
        f = open(path+'Collected_item.json')
        self.data = json.loads(f.read())
        f.close()

class Buying_item:
    def insert_data(self,l,bill_no):
        for i in l:
            l1=[]
            d4={}
            d4["bill_no"]=bill_no
            d4["item_name"]=i["item_name"]
            d4["item_weight"]=i["item_weight"]
            d4["item_price"]=i["item_price"]
            d4["item_pricepergram"]=i["item_pricepergram"]
            l1.append(d4)
            self.data[0]["Buying_item"].append(l1)
        f=open(path+'Buying_item.json','w')
        f.write(json.dumps(self.data))
        f.close()
    
    def __init__(self):
        f = open(path+'Buying_item.json')
        self.data = json.loads(f.read())
        f.close()
class Bill:
    def insert_data(self,bill_no,date,cid):
        l=[]
        d={}
        id1="0"
        d["bill_no"]=bill_no
        d["date"]=str(date)
        d["cid"]=cid
        d["location"]=path+str(bill_no)+".html"
        l.append(d)
        self.data[0]["bill"].append(l)
        f=open(path+'bill.json','w')
        f.write(json.dumps(self.data))
        f.close()
        print(self.data)
        
    def get_new_bill_no(self):
        if len(self.data[0]["bill"])>0:
            return  str(int(self.data[0]["bill"][-1][0]["bill_no"])+1)
            #return str(int(float((self.data[0]["bill"][-1][0]["bill_no"]))+1)
        else:
            return "1"
    def delete_data(self,bill_no):
        for i in self.data[0]["bill"]:
                if i[0]["bill_no"]==str(bill_no):
                    self.data[0]["bill"].remove(i)
                    os.remove(path+str(bill_no)+".html")
                    f=open(path+'bill.json','w')
                    f.write(json.dumps(self.data))
                    print(self.data)
                    f.close()    
    def __init__(self):
        f = open(path+'bill.json')
        self.data = json.loads(f.read())
        f.close()

class Customer_detail:
    def insert_data(self,name,add,pending_amount,bill_no,mob_no):
        l=[]
        d={}
        found=0
        id1="0"
        for i in self.data[0]["customer_detail"]:
            #print(i[0])
            if name==i[0]["name"]:
                found=1
                id1=i[0]["id"]
                pending_amount=str(float(pending_amount))
                i[0]["name"]=name
                i[0]["add"]=add
                i[0]["pending_amount"]=pending_amount
                i[0]["bill no's list"]+=","+bill_no
                i[0]["mob_no"]=mob_no
                break
            else:
                id1=i[0]["id"] # last id store
        if found==0:
            id1=float(id1)+1
            d["id"]=id1
            d["name"]=name
            d["add"]=add
            d["pending_amount"]=pending_amount
            d["bill no's list"]=bill_no
            d["mob_no"]=mob_no
            l.append(d)
            self.data[0]["customer_detail"].append(l)
        f=open(path+'customer_detail.json','w')
        f.write(json.dumps(self.data))
        #prfloat(self.data)
        f.close()
        return id1
    def update_data(self,name):
        row=float(ex.customer_detail.customer_table.currentRow())
        for i in self.data[0]["customer_detail"]:
            if name==i[0]["name"]:
                id1=i[0]["id"]
                i[0]["add"]=ex.customer_detail.customer_table.item(row,1).text()
                i[0]["pending_amount"]=ex.customer_detail.customer_table.item(row,2).text()
                i[0]["mob_no"]=ex.customer_detail.customer_table.item(row,3).text()
                f=open(path+'customer_detail.json','w')
                f.write(json.dumps(self.data))
                f.close()
                return True
        return True
                
    #get data
    def get_data(self,name):
        for i in self.data[0]["customer_detail"]:
            if name==i[0]["name"]:
                return i[0]
        return None
    
    def get_all_data(self):
        return self.data[0]["customer_detail"]    
    
    
    def delete_data(self,name):
        for i in self.data[0]["customer_detail"]:
            if i[0]["name"]==name:
                print(name)
                self.data[0]["customer_detail"].remove(i)
                x=i[0]["bill no's list"].split(",")
                try:
                    for k in x:
                        Bill().delete_data(k)
                except:
                    pass
                f=open(path+'customer_detail.json','w')
                f.write(json.dumps(self.data))
                print(self.data)
                f.close()
                return True
        return False
        
    def __init__(self):
        f = open(path+"customer_detail.json")
        self.data = json.loads(f.read())
        f.close()

         
class font_class(QFont):
    def __init__(self):
        super().__init__("Times New Roman",12)

#  billing page class
class Billing(QWidget):
    def item_weight_validator(self,name,weight,call):
        if call=="add":
            for i in self.item_validator_list:
                if i[0]["item_name"]==name:
                    #print(i[0]["item_weight"])
                    if float(float(i[0]["item_weight"])-float(weight))>=0.0:
                        i[0]["item_weight"]=str(float(i[0]["item_weight"])-float(weight))
                        print(i[0]["item_weight"])
                        return True
                    else:
                        print(i[0]["item_weight"])
                        return False
                    
        else:
            for i in self.item_validator_list:
                if i[0]["item_name"]==name:
                    i[0]["item_weight"]=str(float(i[0]["item_weight"])+float(weight))
                    print(i[0]["item_weight"])

                              
    def item_list_load(self):
        s=Stock_detail()
        temp=s.get_all_data()
        for i  in temp:
            self.item_validator_list.append(i)
        print(self.item_validator_list)
    def weight_update(self):
        print("update weight")
        print(self.item_validator_list)
        f = open(path+"stock_detail.json")
        self.data1 = json.loads(f.read())
        f.close()
        self.data1[0]["stock_detail"].clear()
        for i in self.item_validator_list:
            self.data1[0]["stock_detail"].append(i)
        f=open(path+"stock_detail.json",'w')
        f.write(json.dumps(self.data1))
        f.close()
    def bill_genrate(self,data,bill_no):
        try:
            loc = (path+"demo.xlsx")
            #print(loc)
            book = load_workbook(loc)
            
            cust_details = 'Name : ' + data['customer_name'] + '     ' + 'Address : ' + data['address'] + '     ' + 'Mo. No. : ' + data['mobile_no']
            book['Jewelry invoice']['A6'].value = cust_details
            book['Jewelry invoice']['D3'].value =data['date']
            '''comment strart book['Jewelry invoice']['A8'].value = 'Name : ' + data['customer_name']
            book['Jewelry invoice']['A9'].value = 'Address : ' + data['address']
            book['Jewelry invoice']['A10'].value = 'Mobile No : ' + data['mobile_no'] comment end'''
            book['Jewelry invoice']['D4'].value = data['bill_no']   
           
            k = 8
            #print(data['item_detail'])
            #print('above')
            for i in data['item_detail']:
                item_name = 'A' + str(k)
                item_weight = 'B' + str(k)
                item_price = 'C' + str(k)
                item_per = 'D' + str(k)
                amount = 'E' + str(k)
                book['Jewelry invoice'][item_name].value = i['item_name']
                book['Jewelry invoice'][item_weight].value = i['item_weight']
                #print('here1')
                book['Jewelry invoice'][item_price].value = i['item_pricepergram']
                #print('here2')
                book['Jewelry invoice'][item_per].value = 'gm'
                #print('here3')
                book['Jewelry invoice'][amount].value = i['item_price']
                k+=1
                #print(len(data['item_detail']) + 7,'  ',k)
            if(self.pan.isChecked()):
                 book['Jewelry invoice']['D2'].value ="ATIPS8371F"
            '''
            book['Jewelry invoice']['E'+str(k+1)].value = data['buying_amount']
            book['Jewelry invoice']['E'+str(k+2)].value = data['pending_amount']
            book['Jewelry invoice']['E'+str(k+3)].value = data['labor_charge']
            book['Jewelry invoice']['E'+str(k+4)].value = data['total']
            book['Jewelry invoice']['E'+str(k+5)].value = data['collected_amount']
            book['Jewelry invoice']['E'+str(k+6)].value = str(float(data['overall_total'])+float(data['labor_charge']))
            #book['Jewelry invoice']['E32'].value = data['new_pending_amount']'''
            styl = Alignment(horizontal='right')  
            styl1 = Font(bold=True, size = 13) 
            #cou = 1
            for i in range(0, 12):
                temp = 'A' + str(i + k) + ':' +'D' + str(i + k)
                book['Jewelry invoice'].merge_cells(temp)
                #print(temp)
                book['Jewelry invoice'].merge_cells(temp)
                #book['Jewelry invoice']['C'+str(i + k)].style = styl
                #book['Jewelry invoice']['C'+str(i + k)].style = styl1
                #cou += 1
            #k=k+1
            m = k
            #book['Jewelry invoice'].merge_cells('C20:D20')
            book['Jewelry invoice']['A'+str(k+1)].value = 'ITEM TOTAL'
            #book['Jewelry invoice']['C'+str(k)].style = 
            book['Jewelry invoice']['A'+str(k+2)].value = 'PENDING AMOUNT'
            #book['Jewelry invoice']['C'+str(k)].style = 
            book['Jewelry invoice']['A'+str(k+3)].value = 'LABOUR CHARGE'
            #book['Jewelry invoice']['C'+str(k)].style = 
            book['Jewelry invoice']['A'+str(k+4)].value = 'TOTAL'
            #book['Jewelry invoice']['C'+str(k)].style = 
            book['Jewelry invoice']['A'+str(k+5)].value = 'COLLECTED AMOUNT'
            #book['Jewelry invoice']['C'+str(k)].style = 
            book['Jewelry invoice']['A'+str(k+6)].value = 'OVERALL TOTAL'
            #book['Jewelry invoice']['C'+str(k)].style = 
            k = m

            book['Jewelry invoice']['E'+str(k+1)].value = data['buying_amount']
            #book['Jewelry invoice']['E'+str(k)].style = 
            book['Jewelry invoice']['E'+str(k+2)].value = data['pending_amount']
            #book['Jewelry invoice']['E'+str(k)].style = 
            book['Jewelry invoice']['E'+str(k+3)].value = data['labor_charge']
            #book['Jewelry invoice']['E'+str(k)].style = 
            book['Jewelry invoice']['E'+str(k+4)].value = data['total']
            #book['Jewelry invoice']['E'+str(k)].style = 
            book['Jewelry invoice']['E'+str(k+5)].value = data['collected_amount']
            #book['Jewelry invoice']['E'+str(k)].style = 
            book['Jewelry invoice']['E'+str(k+6)].value =str(float(data['overall_total'])+float(data['labor_charge']))
            #book['Jewelry invoice']['E'+str(k)].style = 
            #book['Jewelry invoice']['E32'].value = data['new_pending_amount']
            book.save(path+'Jewelry_invoice.xlsx')
            #xlsx to html convertor
            xlsx2html(path+'Jewelry_invoice.xlsx', path+str(bill_no)+".html")
            print("Success..")
            os.startfile(path+str(bill_no)+".html")
            return True
        except:
            return False
    def submit(self):
        # create a detail dicitionary for genrate bill and store in database
        if self.c_name_text.currentText()!="" and  self.collectd_amount_text.text()!="" and self.labor_charge_text.text()!="":
            d={}
            d["customer_name"]=self.c_name_text.currentText()
            d["mobile_no"]=self.c_mobileno_text.text()
            d["address"]=self.c_address_text.text()
            d["bill_no"]=self.c_billno_text.text()
            d["date"]=self.c_date_text.text()
            d["labor_charge"]=str(self.labor_charge_text.text())
            item_detail=[]
            nrows = self.item_table.rowCount()
            for row in range(nrows):
                d1={}
                d1["item_name"]=str(self.item_table.item(row,0).text())
                d1["item_weight"]=str(self.item_table.item(row,1).text())
                d1["item_pricepergram"]=str(self.item_table.item(row,2).text())
                d1["item_price"]=str(self.item_table.item(row,3).text())
                item_detail.append(d1)
            d["item_detail"]=item_detail
            received_item_detail=[]
            nrows = self.received_item_table.rowCount()
            for row in range(nrows):
                d2={}
                d2["received_item_name"]=str(self.received_item_table.item(row,0).text())
                d2["received_item_weight"]=str(self.received_item_table.item(row,1).text())
                d2["received_item_pricepergram"]=str(self.received_item_table.item(row,2).text())
                d2["received_item_price"]=str(self.received_item_table.item(row,3).text())
                received_item_detail.append(d2)
            d["received_item_detail"]=received_item_detail
            d["pending_amount"]=str(self.pending_amount)
            d["buying_amount"]=str(self.total_price)
            d["collected_amount"]=str(self.received_total_price)
            d["overall_total"]=str(self.overall_total)
            #d["collected_amount"]=str(self.collectd_amount_text.text())
            d["new_pending_amount"]=str(self.overall_total-float(self.collectd_amount_text.text())+float(self.labor_charge_text.text()))
            print(d["new_pending_amount"])
            d['total'] = str(self.pending_amount + self.total_price + float(self.labor_charge_text.text()))

            #data base load start to entery
            #Customer_detail fill
            print("hello")
            if (self.bill_genrate(d,d["bill_no"])):
                c=Customer_detail()
                cid=c.insert_data(d["customer_name"],d["address"],d["new_pending_amount"],d["bill_no"],d["mobile_no"])
                print("Update Customer")
                b=Bill()
                b.insert_data(d["bill_no"],d["date"],cid)
                by=Buying_item()
                by.insert_data(d["item_detail"],d["bill_no"])
                ci=Collected_item()
                ci.insert_data(d["received_item_detail"],d["bill_no"])
                self.weight_update()
                #refresh Gui
                #Step 1 clear
                refresh()
                self.c_billno_text.setText(str(int(d["bill_no"])+1))
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Customer name or Collected Ammount Required..")
            msg.setWindowTitle("MessageBox")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
    def calculated(self):
        self.overall_total = (self.total_price-self.received_total_price)+self.pending_amount
        #self.overall_total += float(self.labor_charge_text.text())
        self.total_amount_value.setHidden(False)
        self.total_amount_value.setText(str(self.overall_total))
        return
        
    def received_cellClick(self):
        self.collected_item_amount_value.setHidden(False)
        row=float(self.received_item_table.currentRow())
        self.received_total_price -= float(self.received_item_table.item(row,3).text())
        self.received_item_table.removeRow(row)
        self.collected_item_amount_value.setText(str(self.received_total_price))
        self.calculated()
        
    def received_add_item_into_table(self, item_name, item_weight, item_pricepergram,item_price):
        if item_name!="" and item_weight!="" and item_price!="" and item_pricepergram!="":
            self.collected_item_amount_value.setHidden(False)
            row = self.received_item_table.rowCount()
            l=[]
            l.append(item_name)
            l.append(item_weight)
            l.append(item_pricepergram)
            l.append(item_price)
            l.append(QPushButton("Delete",self))
            self.received_item_table.setRowCount(row+1)
            col = 0
            for i in l:
                self.received_item_table.setVerticalHeaderLabels( str(i + 1) for i in range(row+1))
                if col == 4:
                    self.received_item_table.setCellWidget(row, 4, l[4])
                    l[4].clicked.connect(self.received_cellClick)
                else:
                    cell = QTableWidgetItem(str(i))
                    cell.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.received_item_table.setItem(row, col, cell)
                col += 1
            self.received_total_price = sum([float(self.received_item_table.item(i,3).text()) for i in range(row + 1)])
            self.collected_item_amount_value.setText(str(self.received_total_price))
            self.calculated()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Item name or item price or weight is required")
            msg.setWindowTitle("MessageBox")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
    
    def calc(self):
        try:
            self.i_price_text.setText(str(float(self.i_pricepergram_text.text())*float(self.i_weight_text.text())))
        except:
            pass
    def receive_calc(self):
        try:
            self.received_i_price_text.setText(str(float(self.received_i_pricepergram_text.text())*float(self.received_i_weight_text.text())))
        except:
            pass

    def cellClick(self):
        row=int(self.item_table.currentRow())
        self.item_weight_validator(self.item_table.item(row,0).text(),self.item_table.item(row,1).text(),"delete")
        self.total_Buying_Ammount_value.setHidden(False)
        self.total_price -= float(self.item_table.item(row,3).text())
        self.item_table.removeRow(row)
        self.total_Buying_Ammount_value.setText(str(self.total_price))
        self.calculated()
        
        
    def add_item_into_table(self, item_name, item_weight, item_pricepergram,item_price):
        if item_name!="" and item_weight!="" and item_price!="" and item_pricepergram!="":
            if self.item_weight_validator(item_name,item_weight,"add"):
                self.total_Buying_Ammount_value.setHidden(False)
                row = self.item_table.rowCount()
                l=[]
                l.append(item_name)
                l.append(item_weight)
                l.append(item_pricepergram)
                l.append(item_price)
                l.append(QPushButton("Delete",self))
                self.item_table.setRowCount(row+1)
                col = 0
                for i in l:
                    self.item_table.setVerticalHeaderLabels("Item Number : " + str(i + 1) for i in range(row+1))
                    if col == 4:
                        self.item_table.setCellWidget(row, 4, l[4])
                        l[4].clicked.connect(self.cellClick)
                    else:
                        #setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
                        cell = QTableWidgetItem(str(i))
                        cell.setFlags(QtCore.Qt.ItemIsEnabled)
                        self.item_table.setItem(row, col, cell)
                    col += 1
                self.total_price = sum([float(self.item_table.item(i,3).text()) for i in range(row + 1)])
                self.total_Buying_Ammount_value.setText(str(self.total_price))
                self.calculated()
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Not suffiecnt Stock")
                msg.setWindowTitle("MessageBox")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Item name or item price or weight is required")
            msg.setWindowTitle("MessageBox")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
    def get_c_name_data(self):
        if self.c_name_text.currentText()=="":
            return 
        data=Customer_detail().get_data(self.c_name_text.currentText())
        #print(data)
        #input("")
        if data !=None:
            self.c_mobileno_text.setText(data["mob_no"])
            self.c_address_text.setText(data["add"])
            self.pending_Ammount_value.setHidden(False)
            self.pending_Ammount_value.setText(data["pending_amount"])
            self.pending_amount=float(data["pending_amount"])
        else:
            self.c_mobileno_text.setText("")
            self.c_address_text.setText("")
            self.pending_Ammount_value.setHidden(False)
            #self.pending_Ammount_value.setText("0")
            self.pending_amount=0
            self.pending_Ammount_value.setText(str(self.pending_amount))
        self.calculated()
    def get_i_name_data(self):
        if self.i_name_text.currentText()=="":
            return
        data=Stock_detail().get_data(self.i_name_text.currentText())
        self.i_weight_text.setText("0")
            
    def __init__(self):
        super().__init__()
        self.item_detail=[]
        self.item_validator_list=[]
        self.item_list_load()
        self.label = QLabel('Customer Detail', self)
        self.label.move(20,0)
        self.label.resize(2000,80)
        self.label.setFont(QFont('Arial', 30))
        self.c_name = QLabel('Customer Name', self)
        self.c_name.setFont(font_class())
        self.c_name.move(25,110)     
        self.c_name_text=QComboBox(self)
        self.c_name_text.setFocusPolicy(Qt.StrongFocus)
        self.c_name_text.setEditable(True)
        self.c_name_text.move(180,110)
        self.c_name_text.setInsertPolicy(QComboBox.NoInsert) 
        self.pFilterModel = QSortFilterProxyModel(self.c_name_text)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.c_name_text.model())
        self.completer = QCompleter(self.pFilterModel, self.c_name_text)
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.c_name_text.setCompleter(self.completer)
        self.c_name_text.lineEdit().textEdited.connect(self.pFilterModel.setFilterFixedString)
        self.l=[]
        self.customer_list=[]
        c=Customer_detail()
        temp=c.get_all_data()
        for i  in temp:
            self.l.append(i[0])
            self.customer_list.append(i[0]["name"])
        self.c_name_text.addItems(self.customer_list)
        self.c_name_text.resize(200,30)
        self.c_name_text.setFont(font_class())
        self.c_name_text.setCurrentText('')
        self.c_name_text.currentTextChanged.connect(self.get_c_name_data)
        self.c_mobileno = QLabel('Mobile No.', self)
        self.c_mobileno.setFont(font_class())
        self.c_mobileno.move(420,110)    
        self.c_mobileno_text=QLineEdit(self)
        self.c_mobileno_text.move(554,110)
        self.c_mobileno_text.resize(200,30)
        self.c_mobileno_text.setFont(font_class())
        
        #label for cusomer address
        
        self.c_address = QLabel('Address.', self)
        self.c_address.setFont(font_class())
        self.c_address.move(794,110) 
        
        #edit line for c_address
        
        self.c_address_text=QLineEdit(self)
        self.c_address_text.move(900,110)
        self.c_address_text.resize(200,30)
        self.c_address_text.setFont(font_class())
        #label for date
        self.c_date = QLabel('Date.', self)
        self.c_date.setFont(font_class())
        self.c_date.move(1140,110) 
        
        self.c_date_text=QLineEdit(self)
        self.c_date_text.move(1220,110)
        self.c_date_text.resize(200,30)
        self.c_date_text.setFont(font_class())
        self.c_date_text.setText(str(date.today()))
        #lebel for c_billno
        
        self.c_billno = QLabel('Billno', self)
        self.c_billno.setFont(font_class())
        self.c_billno.move(1460,110)
        
        #edit line for c_billno
        
        self.c_billno_text=QLineEdit(self)
        self.c_billno_text.move(1540,110)
        self.c_billno_text.resize(100,30)
        self.c_billno_text.setFont(font_class())
        onlyInt = QIntValidator()
        self.c_billno_text.setValidator(onlyInt)
        self.c_billno_text.setEnabled(False)
        self.c_billno_text.setText(str(Bill().get_new_bill_no()))
        #panCard CheckBox
        self.pan_label= QLabel('PAN', self)
        self.pan_label.setFont(font_class())
        self.pan_label.move(1660,110)
        self.pan=QCheckBox(self)
        self.pan.move(1720,110)
        self.pan.resize(100,30)
        #items detail started
        self.item_detail_label = QLabel('Item Detail', self)
        self.item_detail_label.move(20,170)
        self.item_detail_label.resize(2000,80)
        self.item_detail_label.setFont(QFont('Arial', 30))
        
    
        #label for an item name-
        
        self.i_name = QLabel('Item Name', self)
        self.i_name.setFont(font_class())
        self.i_name.move(25,282) 
        
        #combo box for item name
        
        self.i_name_text=QComboBox(self)
        self.i_name_text.setFocusPolicy(Qt.StrongFocus)
        self.i_name_text.setEditable(True)
        self.i_name_text.move(180,282)
        self.i_name_text.setInsertPolicy(QComboBox.NoInsert) 
        self.pFilterModel = QSortFilterProxyModel(self.i_name_text)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.i_name_text.model())
        self.completer = QCompleter(self.pFilterModel, self.i_name_text)
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.i_name_text.setCompleter(self.completer)
        self.i_name_text.lineEdit().textEdited.connect(self.pFilterModel.setFilterFixedString)
        self.item_list=[]
        s=Stock_detail()
        temp=s.get_all_data()
        for i  in temp:
            self.l.append(i[0])
            self.item_list.append(i[0]["item_name"])
        self.item_list=list(set(self.item_list))
        print(self.item_list)
        self.i_name_text.addItems(self.item_list)
        self.i_name_text.resize(200,30)
        self.i_name_text.setFont(font_class())
        self.i_name_text.setCurrentText('')
        self.i_name_text.currentTextChanged.connect(self.get_i_name_data)
        
        #weight label for an item
        
        self.i_weight = QLabel('Item Weight', self)
        self.i_weight.setFont(font_class())
        self.i_weight.move(420,282)
        
        #weight combo box for an item
        self.i_weight_text=QLineEdit(self)
        self.i_weight_text.move(555,282)
        self.i_weight_text.resize(200,30)
        self.i_weight_text.setFont(font_class())
        onlyInt = QDoubleValidator()
        self.i_weight_text.setText("0")
        self.i_weight_text.setValidator(onlyInt)
        self.i_weight_text.setFont(font_class())
        self.i_weight_text.textChanged.connect(self.calc)
        #label for an item price
        
        self.i_pricepergram = QLabel('PricePerGram.', self)
        self.i_pricepergram.setFont(font_class())
        self.i_pricepergram.move(795,282)
        
        #edit line for an item price
        
        self.i_pricepergram_text=QLineEdit(self)
        self.i_pricepergram_text.move(910,282)
        self.i_pricepergram_text.resize(200,30)
        self.i_pricepergram_text.setFont(font_class())
        self.i_pricepergram_text.setText("0")
        onlyInt = QDoubleValidator()
        self.i_pricepergram_text.setFont(font_class())
        self.i_pricepergram_text.textChanged.connect(self.calc)
        self.i_price = QLabel('Price.', self)
        self.i_price.setFont(font_class())
        self.i_price.move(1150,282)
        
        #edit line for an item price
        
        self.i_price_text=QLineEdit(self)
        self.i_price_text.move(1250,282)
        self.i_price_text.resize(200,30)
        self.i_price_text.setFont(font_class())
        onlyInt = QDoubleValidator()
        self.i_price_text.setEnabled(False)
        self.i_price_text.setValidator(onlyInt)
        self.i_price_text.setFont(font_class())
        self.i_price_text.setText("0")
        
        
        self.add=QPushButton("Add Item",self)
        self.add.move(1520,282)
        self.add.resize(100,30)
        #self.add.setStyleSheet("border :4px solid ")
        self.add.clicked.connect(lambda:self.add_item_into_table(self.i_name_text.currentText(), self.i_weight_text.text(), self.i_pricepergram_text.text(), self.i_price_text.text()))
        #table for item
        self.item_table = QTableWidget(self)
        #self.tableWidget.setRowCount(10)
        self.item_table.setColumnCount(5)
        self.item_table.setColumnWidth(0,300)
        self.item_table.setColumnWidth(2,200)
        self.item_table.setFont(font_class())
        self.item_table.setHorizontalHeaderLabels(["Item Name", "Item Weight", "Item Pricepergram",'Item Price',''])
        self.item_table.move(25,340)
        self.item_table.resize(1030,195)
        self.item_table.setEnabled(True)
        self.item_table.setShowGrid(False)
        

        #pending amount label        
        self.pending_Ammount=QLabel("Pending Amount : ",self)
        self.pending_Ammount.setFont(font_class())
        self.pending_Ammount.move(1080,370)
        
        #pending amount label        
        self.pending_amount = 0.0
        self.pending_Ammount_value=QLabel("00000000000000",self)
        self.pending_Ammount_value.setHidden(True)
        self.pending_Ammount_value.setFont(font_class())
        self.pending_Ammount_value.move(1230,370)
        
        
        #buying amount label
        self.total_Buying_Ammount=QLabel("Buying Amount  : ",self)
        self.total_Buying_Ammount.setFont(font_class())
        self.total_Buying_Ammount.move(1080,410)
        
        #buying amount label
        self.total_price=0.0
        #self.buying_amount=0
        self.total_Buying_Ammount_value=QLabel("0000000000000",self)
        self.total_Buying_Ammount_value.setHidden(True)
        self.total_Buying_Ammount_value.setFont(font_class())
        self.total_Buying_Ammount_value.move(1230,410)
        
        
        #customer received item
   
        #items detail started
        self.received_item_detail_label = QLabel('Received Item Detail', self)
        self.received_item_detail_label.move(20,560)
        self.received_item_detail_label.resize(2000,80)
        #self.label.setStyleSheet("border: 1px solid black; color:gold; background-color:black")
        self.received_item_detail_label.setFont(QFont('Arial', 30))
        
    
        #label for an item name
        
        self.received_i_name = QLabel('Item Name', self)
        self.received_i_name.setFont(font_class())
        self.received_i_name.move(25,672) 
        
        #combo box for item name
        
        self.received_i_name_text=QComboBox(self)
        self.received_i_name_text.setFocusPolicy(Qt.StrongFocus)
        self.received_i_name_text.setEditable(True)
        self.received_i_name_text.move(180,672)
        self.received_i_name_text.setInsertPolicy(QComboBox.NoInsert) 
        self.pFilterModel = QSortFilterProxyModel(self.received_i_name_text)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.received_i_name_text.model())
        self.completer = QCompleter(self.pFilterModel, self.received_i_name_text)
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.received_i_name_text.setCompleter(self.completer)
        self.received_i_name_text.lineEdit().textEdited.connect(self.pFilterModel.setFilterFixedString)
        string_list = []#data base load
        self.received_i_name_text.addItems(string_list)
        self.received_i_name_text.resize(200,30)
        self.received_i_name_text.setFont(font_class())
        self.received_i_name_text.setCurrentText('')
        
        #weight label for an item
        
        self.received_i_weight = QLabel('Item Weight', self)
        self.received_i_weight.setFont(font_class())
        self.received_i_weight.move(420,672)
        
        #weight combo box for an item
        #weight combo box for an item
        self.received_i_weight_text=QLineEdit(self)
        self.received_i_weight_text.move(555,672)
        self.received_i_weight_text.setFont(font_class())
        onlyInt = QDoubleValidator()
        self.received_i_weight_text.setValidator(onlyInt)
        self.received_i_weight_text.setText("0")
        self.received_i_weight_text.textChanged.connect(self.receive_calc)
        #label for an item price
        
        self.received_i_pricepergram = QLabel('Pricepegram.', self)
        self.received_i_pricepergram.setFont(font_class())
        self.received_i_pricepergram.move(795,672)
        
        #edit line for an item price
        
        self.received_i_pricepergram_text=QLineEdit(self)
        self.received_i_pricepergram_text.move(900,672)
        self.received_i_pricepergram_text.resize(200,30)
        self.received_i_pricepergram_text.setFont(font_class())
        self.received_i_pricepergram_text.setValidator(onlyInt)
        self.received_i_pricepergram_text.setText("0")
        self.received_i_pricepergram_text.textChanged.connect(self.receive_calc)
        self.received_i_price = QLabel('Price.', self)
        self.received_i_price.setFont(font_class())
       
        self.received_i_price.move(1140,672)
        
        #edit line for an item price
        
        self.received_i_price_text=QLineEdit(self)
        self.received_i_price_text.move(1280,672)
        self.received_i_price_text.resize(200,30)
        self.received_i_price_text.setText("0")
        self.received_i_price_text.setEnabled(False)
        self.received_i_price_text.setFont(font_class())
        
        self.received_add=QPushButton("Add Item",self)
        self.received_add.move(1520,672)
        self.received_add.resize(100,30)
        #self.add.setStyleSheet("border :4px solid ")
        self.received_add.clicked.connect(lambda:self.received_add_item_into_table(self.received_i_name_text.currentText(), self.received_i_weight_text.text(), self.received_i_pricepergram_text.text(), self.received_i_price_text.text()))
        #table for item
        self.received_item_table = QTableWidget(self)
        #self.tableWidget.setRowCount(10)
        self.received_item_table.setColumnCount(5)
        self.received_item_table.setColumnWidth(0,300)
        self.received_item_table.setColumnWidth(2,200)
        self.received_item_table.setFont(font_class())
        self.received_item_table.setHorizontalHeaderLabels(["Item Name", "Item Weight", "Item Pricepergram",'Item Price',''])
        self.received_item_table.move(25,730)
        self.received_item_table.resize(950,195)
        self.received_item_table.setEnabled(True)
        self.received_item_table.setShowGrid(False)
        

        #Collected Item Amount label        
        
        self.collected_item_amount=QLabel("Collected Item Amount : ",self)
        self.collected_item_amount.setFont(font_class())
        self.collected_item_amount.move(1050,760)
        
        #collected item value label
        self.received_total_price = 0
        self.collected_item_amount_value=QLabel('00000000000000000000',self)
        self.collected_item_amount_value.setHidden(True)
        self.collected_item_amount_value.setFont(font_class())
        self.collected_item_amount_value.move(1260,760)
        #self.collected_item_amount_value.setText("0")
        self.collected_item_amount_value.setHidden(True)
        
        #total amount
        self.total_amount=QLabel("Total Amount  : ",self)
        self.total_amount.setFont(font_class())
        self.total_amount.move(1050,810)
        
        #total amount value
        self.overall_total = 0
        self.total_amount_value=QLabel("000000000000000",self)
        self.total_amount_value.setHidden(True)
        self.total_amount_value.setFont(font_class())
        self.total_amount_value.move(1180,810)
        
        #collected amount label
        self.collectd_amount=QLabel("Collected Amount : ",self)
        self.collectd_amount.setFont(font_class())
        self.collectd_amount.move(1050,900)
        
        #collected amount line edit
        self.collectd_amount_text=QLineEdit(self)
        self.collectd_amount_text.setFont(font_class())
        self.collectd_amount_text.setValidator(onlyInt)
        self.collectd_amount_text.setText("0")
        self.collectd_amount_text.move(1240,900)
        
        self.labor_charge=QLabel("labor charge : ",self)
        self.labor_charge.setFont(font_class())
        self.labor_charge.move(1460,900)
        
        #collected amount line edit
        self.labor_charge_text=QLineEdit(self)
        self.labor_charge_text.setFont(font_class())
        self.labor_charge_text.setValidator(onlyInt)
        self.labor_charge_text.move(1590,900)
        self.labor_charge_text.setText("0")
        self.genrate_bill=QPushButton("Genrate Bill",self)
        self.genrate_bill.move(1800,900)
        self.genrate_bill.clicked.connect(self.submit)

# add stock
class AddStock(QWidget):
    
    def get_i_name_data(self):
        if self.i_name_text.currentText()=="":
            return
        data=Stock_detail().get_data(self.i_name_text.currentText())
        self.i_weight_text.setText("0")
    
    def submit(self):
        s=Stock_detail()
        row=0
        nrows = self.item_table.rowCount()
        for row in range(nrows):
            s.insert_data(self.item_table.item(row,0).text(),self.item_table.item(row,1).text())
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("Stock Updated")
        msg.setWindowTitle("Update Stock")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        retval=msg.exec_()
        refresh()
        
    def cellClick(self):
        row=int(self.item_table.currentRow())
        self.item_table.removeRow(row)
        
    def add_item_into_table(self, item_name, item_weight):
        if item_name!="" and item_weight!="":
            row = self.item_table.rowCount()
            l=[]
            l.append(item_name)
            l.append(item_weight)
            l.append(QPushButton("Delete",self))
            self.item_table.setRowCount(row+1)
            col = 0
            for i in l:
                self.item_table.setVerticalHeaderLabels("Item Number : " + str(i + 1) for i in range(row+1))
                if col == 2:
                    self.item_table.setCellWidget(row, 2, l[2])
                    l[2].clicked.connect(self.cellClick)
                else:
                    cell = QTableWidgetItem(str(i))
                    cell.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.item_table.setItem(row, col, cell)
                col += 1
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Item name and item price and weight is required")
            msg.setWindowTitle("MessageBox")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
    def __init__(self):
        super().__init__()
        self.l=[]
        self.label = QLabel('Raghunandan Jewellers', self)
        self.label.move(0,0)
        self.label.resize(2000,100)
        self.label.setStyleSheet("border: 1px solid black; color:gold; background-color:black")
        self.label.setFont(QFont('Arial', 30))
        #add stock label
        self.label = QLabel('Add Stock', self)
        self.label.move(0,110)
        self.label.resize(1000,100)
        self.label.setFont(QFont('Arial', 30))
        self.i_name = QLabel('Item Name', self)
        self.i_name.setFont(font_class())
        self.i_name.move(25,282) 
        
        #combo box for item name
        
        self.i_name_text=QComboBox(self)
        self.i_name_text.setFocusPolicy(Qt.StrongFocus)
        self.i_name_text.setEditable(True)
        self.i_name_text.move(180,282)
        self.i_name_text.setInsertPolicy(QComboBox.NoInsert) 
        self.pFilterModel = QSortFilterProxyModel(self.i_name_text)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.i_name_text.model())
        self.completer = QCompleter(self.pFilterModel, self.i_name_text)
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.i_name_text.setCompleter(self.completer)
        self.i_name_text.lineEdit().textEdited.connect(self.pFilterModel.setFilterFixedString)
        self.item_list=[]
        s=Stock_detail()
        temp=s.get_all_data()
        for i  in temp:
            self.l.append(i[0])
            self.item_list.append(i[0]["item_name"])
        self.item_list=list(set(self.item_list))
        self.i_name_text.addItems(self.item_list)
        self.i_name_text.resize(200,30)
        self.i_name_text.setFont(font_class())
        self.i_name_text.setCurrentText('')
        self.i_name_text.currentTextChanged.connect(self.get_i_name_data)
        
        #weight label for an item
        
        self.i_weight = QLabel('Item Weight', self)
        self.i_weight.setFont(font_class())
        self.i_weight.move(420,282)
        
        #weight combo box for an item
        
        self.i_weight_text=QLineEdit(self)
        self.i_weight_text.move(555,282)
        self.i_weight_text.resize(200,30)
        self.i_weight_text.setFont(font_class())
        onlyInt = QDoubleValidator()
        self.i_weight_text.setValidator(onlyInt)
        self.i_weight_text.setFont(font_class())
        self.add=QPushButton("Add Item",self)
        self.add.move(900,282)
        self.add.resize(100,30)
        #self.add.setStyleSheet("border :4px solid ")
        self.add.clicked.connect(lambda:self.add_item_into_table(self.i_name_text.currentText(), self.i_weight_text.text()))
        #table for item
        self.item_table = QTableWidget(self)
        #self.tableWidget.setRowCount(10)
        self.item_table.setColumnCount(4)
        self.item_table.setColumnWidth(0,300)
        self.item_table.setColumnWidth(1,300)
        self.item_table.setColumnWidth(2,300)
        self.item_table.setColumnWidth(3,40)
        self.item_table.setFont(font_class())
        self.item_table.setHorizontalHeaderLabels(["Item Name", "Item Weight", "Item Price",''])
        self.item_table.move(25,340)
        self.item_table.resize(1200,195)
        self.item_table.setEnabled(True)
        self.item_table.setShowGrid(False)
        self.submit_button=QPushButton("submit",self)
        self.submit_button.move(1000,800)
        self.submit_button.clicked.connect(self.submit)
        

#stock detail
class StockDetail(QWidget):
    def update_data(self):
        row=int(self.stock_table.currentRow())
        name=self.stock_table.item(row,0).text()
        weight=self.stock_table.item(row,1).text()
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        row=int(self.stock_table.currentRow())
        msg.setText("Want to Update "+self.stock_table.item(row,0).text())
        msg.setWindowTitle("Update Stock")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        retval=msg.exec_()
        if retval==QMessageBox.Ok:
            if Stock_detail().update_data(name,weight):
                refresh()
        else:
            refresh()
    def cellClick(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        row=int(self.stock_table.currentRow())
        name=self.stock_table.item(row,0).text()
        weight=self.stock_table.item(row,1).text()
        msg.setText("Want to Delete"+self.stock_table.item(row,0).text())
        msg.setWindowTitle("DeleteStock")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        retval=msg.exec_()
        if retval==QMessageBox.Ok:
            if Stock_detail().delete_data(name):
                refresh()       
    def __init__(self):
        super().__init__()
        self.label = QLabel('Raghunandan Jewellers', self)
        self.label.move(0,0)
        self.label.resize(2000,100)
        #self.label.setStyleSheet("border: 1px solid black; color:gold; background-color:black")
        self.label.setFont(QFont('Arial', 30))
        self.stock_table = QTableWidget(self)
        self.stock_table.setColumnCount(4)
        self.stock_table.setColumnWidth(0,800)
        self.stock_table.setColumnWidth(1,800)
        self.stock_table.setColumnWidth(2,100)
        self.stock_table.setColumnWidth(3,100)
        self.stock_table.setHorizontalHeaderLabels(["Item_Name", "Item_weight","",""])
        self.stock_table.move(0,120)
        self.stock_table.setEnabled(True)
        self.stock_table.resize(2100,900)
        c=Stock_detail()
        data=c.get_all_data()
        row = 0
        for i in data:
             self.stock_table.setRowCount(row+1)
             l=[]
             l.append(i[0]["item_name"])
             l.append(i[0]["item_weight"])
             l.append(QPushButton("Delete",self))
             l.append(QPushButton("Update",self))
             col = 0
             for i in l:
                 if col == 2:
                    self.stock_table.setCellWidget(row, 2, l[2])
                    l[2].clicked.connect(self.cellClick)
                 elif col==3:
                     self.stock_table.setCellWidget(row, 3, l[3])
                     l[3].clicked.connect(self.update_data)
                 else:
                    cell = QTableWidgetItem(str(i))
                    self.stock_table.setItem(row, col, cell)
                 col += 1
             row+=1
# customer detail
class CustomerDetail(QWidget):
    def update_data(self):
        row=int(self.customer_table.currentRow())
        name=self.customer_table.item(row,0).text()
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        row=int(self.customer_table.currentRow())
        msg.setText("Want to Update "+self.customer_table.item(row,0).text())
        msg.setWindowTitle("Update Customer")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        retval=msg.exec_()
        if retval==QMessageBox.Ok:
            if Customer_detail().update_data(name):
                print("True")
                refresh()
        else:
            refresh()  
    def cellClick(self):
        print("hello")
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        row=int(self.customer_table.currentRow())
        msg.setText("Want to Delete"+self.customer_table.item(row,0).text())
        msg.setWindowTitle("DeleteCustomer")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        retval=msg.exec_()
        if retval==QMessageBox.Ok:
            #print("in delete")
            print(self.customer_table.item(row,0).text())
            if Customer_detail().delete_data(self.customer_table.item(row,0).text()):
                refresh() 
            else:
                refresh()     
    def __init__(self):
        super().__init__()
        self.label = QLabel('Raghunandan Jewellers', self)
        self.label.move(0,0)
        self.label.resize(2000,100)
        #self.label.setStyleSheet("border: 1px solid black; color:gold; background-color:black")
        self.label.setFont(QFont('Arial', 30))
        self.customer_table = QTableWidget(self)
        self.customer_table.setColumnCount(7)
        self.customer_table.setColumnWidth(0,400)
        self.customer_table.setColumnWidth(1,200)
        self.customer_table.setColumnWidth(2,200)
        self.customer_table.setColumnWidth(3,200)
        self.customer_table.setColumnWidth(4,500)
        self.customer_table.setColumnWidth(5,100)
        self.customer_table.setColumnWidth(6,100)
        self.customer_table.setHorizontalHeaderLabels(["Customer_Name", "Address","Pending Amount","Mobile_no","Bill_no's","",""])
        self.customer_table.move(0,120)
        self.customer_table.setEnabled(True)
        self.customer_table.resize(2000,900)
        c=Customer_detail()
        data=c.get_all_data()
        row = 0
        for i in data:
             self.customer_table.setRowCount(row+1)
             l=[]
             l.append(i[0]["name"])
             l.append(i[0]["add"])
             l.append(i[0]["pending_amount"])
             l.append(i[0]["mob_no"])
             l.append(i[0]["bill no's list"])
             l.append(QPushButton("Delete",self))
             l.append(QPushButton("Update",self))
             col = 0
             for i in l:
                 if col == 5:
                    self.customer_table.setCellWidget(row, 5, l[5])
                    l[5].clicked.connect(self.cellClick)
                 elif col==6:
                     self.customer_table.setCellWidget(row, 6, l[6])
                     l[6].clicked.connect(self.update_data)
                 else:
                    cell = QTableWidgetItem(str(i))
                    self.customer_table.setItem(row, col, cell)
                 col += 1
             row+=1
#main class
class Soni(QMainWindow):
    def __init__(self):
        super().__init__()
        x=0
        y=0
        width=2000
        height=1000
        self.setGeometry(x,y,width,height)#position_X,position_Y,Width,Height
        self.tabs = QTabWidget(self)
        self.billing = Billing()
        self.add_stock = AddStock()
        self.stock_detail = StockDetail()
        self.customer_detail = CustomerDetail()
        self.tabs.resize(2000,1000)
        # Add tabs
        self.tabs.addTab(self.billing,"Billing")
        self.tabs.addTab(self.add_stock,"Add Stock")
        self.tabs.addTab(self.stock_detail,"Stock Detail")
        self.tabs.addTab(self.customer_detail,"Customer Detail")
        self.setWindowTitle("Jwallery")
        self.show()
app = QApplication(sys.argv)
ex=Soni()
sys.exit(app.exec_())